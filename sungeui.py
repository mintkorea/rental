import streamlit as st
import requests
import pandas as pd
from datetime import datetime, timedelta
import pytz
import io
# 엑셀 스타일 자동화를 위한 라이브러리
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import expand_dataframe_to_values

# 1. 초기 설정
st.set_page_config(page_title="성의교정 대관 조회", layout="wide")
KST = pytz.timezone('Asia/Seoul')
now_today = datetime.now(KST).date()

BUILDING_ORDER = ["성의회관", "의생명산업연구원", "옴니버스 파크", "옴니버스파크 의과대학", "옴니버스파크 간호대학", "대학본관", "서울성모별관"]
DEFAULT_BUILDINGS = ["성의회관", "의생명산업연구원"]

# 2. 데이터 로드 함수 (기존 유지)
@st.cache_data(ttl=60)
def get_data(s_date, e_date):
    url = "https://songeui.catholic.ac.kr/ko/service/application-for-rental_calendar.do"
    params = {"mode": "getReservedData", "start": s_date.isoformat(), "end": e_date.isoformat()}
    try:
        res = requests.get(url, params=params, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        raw = res.json().get('res', [])
        rows = []
        for item in raw:
            if not item.get('startDt'): continue
            s_dt = datetime.strptime(item['startDt'], '%Y-%m-%d').date()
            e_dt = datetime.strptime(item['endDt'], '%Y-%m-%d').date()
            allow_day_raw = str(item.get('allowDay', ''))
            allowed_days = [int(d.strip()) for d in allow_day_raw.split(',') if d.strip().isdigit()] if allow_day_raw and allow_day_raw.lower() != 'none' else []
            curr = s_dt
            while curr <= e_dt:
                if s_date <= curr <= e_date:
                    if not allowed_days or (curr.weekday() + 1) in allowed_days:
                        rows.append({
                            '날짜': curr.strftime('%Y-%m-%d'),
                            '요일': ['월','화','수','목','금','토','일'][curr.weekday()],
                            '건물명': str(item.get('buNm', '')).strip(),
                            '장소': item.get('placeNm', '') or '', 
                            '시간': f"{item.get('startTime', '')}~{item.get('endTime', '')}",
                            '행사명': item.get('eventNm', '') or '', 
                            '인원': str(item.get('peopleCount', '')) if item.get('peopleCount') else '-',
                            '부서': item.get('mgDeptNm', '') or '-',
                            '상태': '확정' if item.get('status') == 'Y' else '대기'
                        })
                curr += timedelta(days=1)
        return pd.DataFrame(rows)
    except: return pd.DataFrame()

# 3. 엑셀 자동화 편집 함수
def create_styled_excel(df, selected_bu):
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "대관현황"

    # 데이터 정렬 (날짜순 -> 건물순)
    df['건물순서'] = df['건물명'].apply(lambda x: BUILDING_ORDER.index(x) if x in BUILDING_ORDER else 99)
    df = df.sort_values(by=['날짜', '건물순서', '시간'])
    
    # 엑셀 스타일 정의
    header_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 헤더 작성
    headers = ['날짜', '요일', '건물명', '장소', '시간', '행사명', '인원', '부서', '상태']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 데이터 작성 및 자동 편집
    for row_num, row_data in enumerate(df[headers].values, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            # 행사명(6번째 열)만 왼쪽 정렬, 나머지는 가운데 정렬
            cell.alignment = left_align if col_num == 6 else center_align

    # 열 너비 자동 조절
    column_widths = [12, 5, 15, 20, 15, 40, 8, 20, 8]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    wb.save(output)
    return output.getvalue()

# 4. UI 로직
st.sidebar.title("📅 설정")
s_date = st.sidebar.date_input("시작일", value=now_today)
e_date = st.sidebar.date_input("종료일", value=s_date)
sel_bu = st.sidebar.multiselect("건물 필터", options=BUILDING_ORDER, default=DEFAULT_BUILDINGS)

df = get_data(s_date, e_date)

# 사이드바 버튼들
with st.sidebar:
    if not df.empty:
        filtered_df = df[df['건물명'].isin(sel_bu)]
        
        # 엑셀 다운로드 버튼
        excel_data = create_styled_excel(filtered_df, sel_bu)
        st.download_button(
            label="📥 엑셀(Excel) 다운로드",
            data=excel_data,
            file_name=f"rental_{s_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# 화면 출력 (기존 유지)
st.markdown('<div class="main-title">🏫 성의교정 대관 현황</div>', unsafe_allow_html=True)
# ... (이하 화면 출력 로직은 기존과 동일)
